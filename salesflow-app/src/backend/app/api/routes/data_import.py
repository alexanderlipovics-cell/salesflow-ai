"""
╔════════════════════════════════════════════════════════════════════════════╗
║  DATA IMPORT API                                                           ║
║  CSV/Excel Import für Bestandskunden                                       ║
╚════════════════════════════════════════════════════════════════════════════╝

Features:
- CSV-Upload mit automatischer Spaltenerkennung
- Excel-Upload (.xlsx, .xls)
- Feldmapping (automatisch + manuell)
- Daten-Validierung
- Duplikat-Erkennung
- Fortschrittsanzeige
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, BackgroundTasks
from pydantic import BaseModel, Field, EmailStr
from typing import Optional, List, Dict, Any
from datetime import datetime
from enum import Enum
import csv
import io
import json
import uuid
import logging

from ...db.deps import get_db, get_current_user, CurrentUser
from ...db.supabase import get_supabase

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/import", tags=["data-import"])


# =============================================================================
# ENUMS & SCHEMAS
# =============================================================================

class ImportStatus(str, Enum):
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    PARTIAL = "partial"


class LeadStatus(str, Enum):
    NEW = "new"
    COLD = "cold"
    WARM = "warm"
    HOT = "hot"
    CUSTOMER = "customer"
    LOST = "lost"


class FieldMapping(BaseModel):
    """Mapping einer CSV-Spalte zu einem Datenbankfeld."""
    csv_column: str
    db_field: str
    transform: Optional[str] = None  # z.B. "uppercase", "phone_format"


class ImportPreviewRow(BaseModel):
    """Eine Vorschau-Zeile des Imports."""
    row_number: int
    data: Dict[str, Any]
    errors: List[str] = []
    warnings: List[str] = []
    is_duplicate: bool = False


class ImportConfig(BaseModel):
    """Import-Konfiguration."""
    field_mappings: List[FieldMapping]
    skip_duplicates: bool = True
    default_status: LeadStatus = LeadStatus.WARM
    default_source: str = "csv_import"
    tags: List[str] = []


class ImportResult(BaseModel):
    """Ergebnis eines Imports."""
    import_id: str
    status: ImportStatus
    total_rows: int
    imported: int
    skipped: int
    errors: int
    duplicates: int
    error_details: List[Dict[str, Any]] = []
    created_at: datetime


class ImportPreviewResponse(BaseModel):
    """Antwort auf Preview-Request."""
    detected_columns: List[str]
    suggested_mappings: Dict[str, str]
    preview_rows: List[ImportPreviewRow]
    total_rows: int
    estimated_duplicates: int


# =============================================================================
# STANDARD-FELDER FÜR LEADS
# =============================================================================

LEAD_DB_FIELDS = {
    "first_name": {"type": "string", "required": False, "aliases": ["vorname", "firstname", "first name", "name"]},
    "last_name": {"type": "string", "required": False, "aliases": ["nachname", "lastname", "last name", "surname"]},
    "name": {"type": "string", "required": False, "aliases": ["vollständiger name", "full name", "fullname", "kontakt"]},
    "email": {"type": "email", "required": False, "aliases": ["e-mail", "emailadresse", "email address", "mail"]},
    "phone": {"type": "phone", "required": False, "aliases": ["telefon", "tel", "handy", "mobile", "phone number", "rufnummer"]},
    "company": {"type": "string", "required": False, "aliases": ["firma", "unternehmen", "company name", "organisation"]},
    "position": {"type": "string", "required": False, "aliases": ["titel", "job", "job title", "rolle", "funktion"]},
    "status": {"type": "enum", "required": False, "aliases": ["lead status", "lead-status", "kategorie"], "enum": LeadStatus},
    "source": {"type": "string", "required": False, "aliases": ["quelle", "herkunft", "lead source", "akquisekanal"]},
    "notes": {"type": "text", "required": False, "aliases": ["notizen", "bemerkungen", "comments", "anmerkungen"]},
    "city": {"type": "string", "required": False, "aliases": ["stadt", "ort", "wohnort"]},
    "country": {"type": "string", "required": False, "aliases": ["land", "country"]},
    "instagram": {"type": "string", "required": False, "aliases": ["instagram handle", "ig", "insta"]},
    "linkedin": {"type": "string", "required": False, "aliases": ["linkedin url", "linkedin profil"]},
    "last_contact": {"type": "date", "required": False, "aliases": ["letzter kontakt", "last contacted", "kontaktdatum"]},
    "tags": {"type": "array", "required": False, "aliases": ["schlagworte", "kategorien", "labels"]},
}


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def detect_delimiter(content: str) -> str:
    """Erkennt den CSV-Delimiter."""
    sample = content[:4000]
    delimiters = [',', ';', '\t', '|']
    counts = {d: sample.count(d) for d in delimiters}
    return max(counts, key=counts.get)


def suggest_field_mapping(csv_columns: List[str]) -> Dict[str, str]:
    """Schlägt automatisches Feldmapping basierend auf Spaltennamen vor."""
    mappings = {}
    
    for col in csv_columns:
        col_lower = col.lower().strip()
        
        for db_field, config in LEAD_DB_FIELDS.items():
            # Exakter Match
            if col_lower == db_field:
                mappings[col] = db_field
                break
            
            # Alias Match
            for alias in config.get("aliases", []):
                if col_lower == alias.lower() or alias.lower() in col_lower:
                    mappings[col] = db_field
                    break
            
            if col in mappings:
                break
    
    return mappings


def validate_row(row: Dict[str, Any], mappings: Dict[str, str]) -> tuple[Dict[str, Any], List[str], List[str]]:
    """Validiert eine Zeile und gibt transformierte Daten + Errors + Warnings zurück."""
    errors = []
    warnings = []
    transformed = {}
    
    for csv_col, db_field in mappings.items():
        value = row.get(csv_col, "").strip() if row.get(csv_col) else ""
        
        if not value:
            continue
        
        field_config = LEAD_DB_FIELDS.get(db_field, {})
        field_type = field_config.get("type", "string")
        
        # Type-specific validation
        if field_type == "email":
            if "@" not in value or "." not in value:
                warnings.append(f"'{value}' sieht nicht wie eine gültige E-Mail aus")
            transformed[db_field] = value.lower()
        
        elif field_type == "phone":
            # Bereinige Telefonnummer
            cleaned = "".join(c for c in value if c.isdigit() or c == "+")
            if len(cleaned) < 6:
                warnings.append(f"Telefonnummer '{value}' scheint zu kurz")
            transformed[db_field] = cleaned
        
        elif field_type == "enum":
            enum_class = field_config.get("enum")
            if enum_class:
                # Versuche Match
                value_lower = value.lower()
                matched = False
                for enum_val in enum_class:
                    if value_lower == enum_val.value.lower() or value_lower in enum_val.value.lower():
                        transformed[db_field] = enum_val.value
                        matched = True
                        break
                if not matched:
                    transformed[db_field] = LeadStatus.WARM.value
                    warnings.append(f"Status '{value}' nicht erkannt, verwende 'warm'")
            else:
                transformed[db_field] = value
        
        elif field_type == "date":
            # Einfache Datumsvalidierung
            transformed[db_field] = value
        
        elif field_type == "array":
            # Tags als kommaseparierte Liste
            transformed[db_field] = [t.strip() for t in value.split(",") if t.strip()]
        
        else:
            transformed[db_field] = value
    
    # Name zusammenbauen wenn nicht vorhanden
    if "name" not in transformed and ("first_name" in transformed or "last_name" in transformed):
        first = transformed.get("first_name", "")
        last = transformed.get("last_name", "")
        transformed["name"] = f"{first} {last}".strip()
    
    # Mindestens ein Identifikator benötigt
    if not any(transformed.get(f) for f in ["name", "email", "phone"]):
        errors.append("Mindestens Name, E-Mail oder Telefon benötigt")
    
    return transformed, errors, warnings


async def check_duplicate(db, user_id: str, email: str = None, phone: str = None, name: str = None) -> bool:
    """Prüft ob ein Lead bereits existiert."""
    try:
        conditions = []
        
        if email:
            result = await db.from_("leads").select("id").eq("user_id", user_id).eq("email", email).limit(1).execute()
            if result.data:
                return True
        
        if phone:
            result = await db.from_("leads").select("id").eq("user_id", user_id).eq("phone", phone).limit(1).execute()
            if result.data:
                return True
        
        return False
    except Exception as e:
        logger.error(f"Duplicate check error: {e}")
        return False


# =============================================================================
# API ENDPOINTS
# =============================================================================

@router.post("/preview", response_model=ImportPreviewResponse)
async def preview_import(
    file: UploadFile = File(...),
    user: CurrentUser = Depends(get_current_user),
    db = Depends(get_db),
):
    """
    Lädt eine CSV/Excel-Datei und zeigt eine Vorschau mit Feldmapping.
    
    Returns:
        - Erkannte Spalten
        - Vorgeschlagenes Mapping
        - Vorschau der ersten 10 Zeilen
        - Geschätzte Duplikate
    """
    try:
        content = await file.read()
        filename = file.filename.lower()
        
        # CSV oder Excel?
        if filename.endswith(('.xlsx', '.xls')):
            # Excel-Import
            try:
                import openpyxl
                from io import BytesIO
                
                wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
                ws = wb.active
                
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    raise HTTPException(status_code=400, detail="Excel-Datei ist leer")
                
                headers = [str(h).strip() if h else f"Spalte_{i}" for i, h in enumerate(rows[0])]
                data_rows = [dict(zip(headers, row)) for row in rows[1:] if any(row)]
                
            except ImportError:
                raise HTTPException(
                    status_code=400, 
                    detail="Excel-Support nicht verfügbar. Bitte als CSV exportieren."
                )
        else:
            # CSV-Import
            try:
                text = content.decode('utf-8')
            except UnicodeDecodeError:
                text = content.decode('latin-1')
            
            delimiter = detect_delimiter(text)
            reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
            headers = reader.fieldnames or []
            data_rows = list(reader)
        
        if not headers:
            raise HTTPException(status_code=400, detail="Keine Spaltenüberschriften gefunden")
        
        if not data_rows:
            raise HTTPException(status_code=400, detail="Keine Daten gefunden")
        
        # Feldmapping vorschlagen
        suggested_mappings = suggest_field_mapping(headers)
        
        # Preview-Zeilen erstellen
        preview_rows = []
        estimated_duplicates = 0
        
        for i, row in enumerate(data_rows[:10]):
            transformed, errors, warnings = validate_row(row, suggested_mappings)
            
            # Duplikat-Check
            is_dup = await check_duplicate(
                db, 
                user.id,
                email=transformed.get("email"),
                phone=transformed.get("phone"),
                name=transformed.get("name")
            )
            if is_dup:
                estimated_duplicates += 1
            
            preview_rows.append(ImportPreviewRow(
                row_number=i + 1,
                data=transformed,
                errors=errors,
                warnings=warnings,
                is_duplicate=is_dup
            ))
        
        # Hochrechnung der Duplikate
        if len(data_rows) > 10 and len(preview_rows) > 0:
            dup_rate = estimated_duplicates / len(preview_rows)
            estimated_duplicates = int(len(data_rows) * dup_rate)
        
        return ImportPreviewResponse(
            detected_columns=headers,
            suggested_mappings=suggested_mappings,
            preview_rows=preview_rows,
            total_rows=len(data_rows),
            estimated_duplicates=estimated_duplicates
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Preview error: {e}")
        raise HTTPException(status_code=500, detail=f"Fehler beim Verarbeiten: {str(e)}")


@router.post("/execute", response_model=ImportResult)
async def execute_import(
    file: UploadFile = File(...),
    config: str = Form(...),  # JSON-String der ImportConfig
    background_tasks: BackgroundTasks = None,
    user: CurrentUser = Depends(get_current_user),
    db = Depends(get_db),
):
    """
    Führt den Import mit der angegebenen Konfiguration aus.
    
    Args:
        file: Die CSV/Excel-Datei
        config: ImportConfig als JSON-String
    
    Returns:
        ImportResult mit Statistiken
    """
    import_id = str(uuid.uuid4())
    
    try:
        # Config parsen
        config_dict = json.loads(config)
        import_config = ImportConfig(**config_dict)
        
        # Datei lesen
        content = await file.read()
        filename = file.filename.lower()
        
        # Parsen
        if filename.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                from io import BytesIO
                
                wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                headers = [str(h).strip() if h else f"Spalte_{i}" for i, h in enumerate(rows[0])]
                data_rows = [dict(zip(headers, row)) for row in rows[1:] if any(row)]
            except ImportError:
                raise HTTPException(status_code=400, detail="Excel-Support nicht verfügbar")
        else:
            try:
                text = content.decode('utf-8')
            except UnicodeDecodeError:
                text = content.decode('latin-1')
            
            delimiter = detect_delimiter(text)
            reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
            data_rows = list(reader)
        
        # Mappings als Dict
        mappings = {m.csv_column: m.db_field for m in import_config.field_mappings}
        
        # Import durchführen
        imported = 0
        skipped = 0
        errors_count = 0
        duplicates = 0
        error_details = []
        
        for i, row in enumerate(data_rows):
            try:
                transformed, row_errors, warnings = validate_row(row, mappings)
                
                if row_errors:
                    errors_count += 1
                    error_details.append({
                        "row": i + 1,
                        "errors": row_errors,
                        "data": {k: str(v)[:50] for k, v in row.items() if v}
                    })
                    continue
                
                # Duplikat-Check
                is_dup = await check_duplicate(
                    db, 
                    user.id,
                    email=transformed.get("email"),
                    phone=transformed.get("phone"),
                    name=transformed.get("name")
                )
                
                if is_dup:
                    duplicates += 1
                    if import_config.skip_duplicates:
                        skipped += 1
                        continue
                
                # Lead erstellen
                lead_data = {
                    "user_id": user.id,
                    "status": transformed.get("status", import_config.default_status.value),
                    "source": transformed.get("source", import_config.default_source),
                    "created_at": datetime.utcnow().isoformat(),
                    "updated_at": datetime.utcnow().isoformat(),
                }
                
                # Felder übertragen
                for field in ["first_name", "last_name", "name", "email", "phone", 
                              "company", "position", "notes", "city", "country",
                              "instagram", "linkedin"]:
                    if field in transformed:
                        lead_data[field] = transformed[field]
                
                # Tags
                tags = transformed.get("tags", []) + import_config.tags
                if tags:
                    lead_data["tags"] = tags
                
                # In DB einfügen
                result = await db.from_("leads").insert(lead_data).execute()
                
                if result.data:
                    imported += 1
                else:
                    errors_count += 1
                    error_details.append({
                        "row": i + 1,
                        "errors": ["Datenbankfehler beim Einfügen"],
                        "data": {k: str(v)[:50] for k, v in transformed.items()}
                    })
                    
            except Exception as e:
                errors_count += 1
                error_details.append({
                    "row": i + 1,
                    "errors": [str(e)],
                    "data": {k: str(v)[:50] for k, v in row.items() if v}
                })
        
        # Status bestimmen
        if errors_count == 0 and imported > 0:
            status = ImportStatus.COMPLETED
        elif imported > 0 and errors_count > 0:
            status = ImportStatus.PARTIAL
        elif imported == 0:
            status = ImportStatus.FAILED
        else:
            status = ImportStatus.COMPLETED
        
        # Import-Log speichern
        try:
            await db.from_("import_logs").insert({
                "id": import_id,
                "user_id": user.id,
                "filename": file.filename,
                "status": status.value,
                "total_rows": len(data_rows),
                "imported": imported,
                "skipped": skipped,
                "errors": errors_count,
                "duplicates": duplicates,
                "created_at": datetime.utcnow().isoformat(),
            }).execute()
        except Exception as e:
            logger.warning(f"Could not log import: {e}")
        
        return ImportResult(
            import_id=import_id,
            status=status,
            total_rows=len(data_rows),
            imported=imported,
            skipped=skipped,
            errors=errors_count,
            duplicates=duplicates,
            error_details=error_details[:20],  # Max 20 Fehler zurückgeben
            created_at=datetime.utcnow()
        )
        
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Ungültige Konfiguration")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Import error: {e}")
        raise HTTPException(status_code=500, detail=f"Import fehlgeschlagen: {str(e)}")


@router.get("/history")
async def get_import_history(
    limit: int = 10,
    user: CurrentUser = Depends(get_current_user),
    db = Depends(get_db),
):
    """Gibt die Import-Historie des Users zurück."""
    try:
        result = await db.from_("import_logs")\
            .select("*")\
            .eq("user_id", user.id)\
            .order("created_at", desc=True)\
            .limit(limit)\
            .execute()
        
        return {"imports": result.data or []}
    except Exception as e:
        logger.error(f"History error: {e}")
        return {"imports": []}


@router.get("/template")
async def download_template():
    """Gibt eine CSV-Vorlage zum Download zurück."""
    headers = [
        "Vorname", "Nachname", "E-Mail", "Telefon", "Firma", 
        "Position", "Status", "Quelle", "Notizen", "Stadt", "Tags"
    ]
    
    example_row = [
        "Max", "Mustermann", "max@example.com", "+49 170 1234567", "Muster GmbH",
        "Geschäftsführer", "warm", "Empfehlung", "Interessiert an Produkt X", "München", "VIP,Entscheider"
    ]
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(headers)
    writer.writerow(example_row)
    
    return {
        "content": output.getvalue(),
        "filename": "leads_import_vorlage.csv",
        "content_type": "text/csv"
    }


@router.post("/quick-import")
async def quick_import(
    file: UploadFile = File(...),
    user: CurrentUser = Depends(get_current_user),
    db = Depends(get_db),
):
    """
    Schneller Import mit automatischem Mapping.
    Für Benutzer die keine manuelle Konfiguration wollen.
    """
    # Erst Preview machen
    await file.seek(0)
    content = await file.read()
    await file.seek(0)
    
    filename = file.filename.lower()
    
    # Parsen
    if filename.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            headers = [str(h).strip() if h else f"Spalte_{i}" for i, h in enumerate(rows[0])]
        except ImportError:
            raise HTTPException(status_code=400, detail="Excel-Support nicht verfügbar")
    else:
        try:
            text = content.decode('utf-8')
        except UnicodeDecodeError:
            text = content.decode('latin-1')
        
        delimiter = detect_delimiter(text)
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        headers = reader.fieldnames or []
    
    # Auto-Mapping
    mappings = suggest_field_mapping(headers)
    
    if not mappings:
        raise HTTPException(
            status_code=400, 
            detail="Keine Spalten erkannt. Bitte verwende Spaltenüberschriften wie 'Name', 'E-Mail', 'Telefon'"
        )
    
    # Config erstellen
    config = ImportConfig(
        field_mappings=[
            FieldMapping(csv_column=csv_col, db_field=db_field)
            for csv_col, db_field in mappings.items()
        ],
        skip_duplicates=True,
        default_status=LeadStatus.WARM,
        default_source="quick_import",
        tags=["importiert"]
    )
    
    # Import ausführen
    await file.seek(0)
    return await execute_import(
        file=file,
        config=config.model_dump_json(),
        user=user,
        db=db
    )

